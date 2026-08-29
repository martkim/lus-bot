import io
import logging

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from src import db

logger = logging.getLogger("passion_mate")

# Excel/CSV 포뮬러 인젝션 방지: 이름 등 자유 입력 필드가 '=', '+', '-', '@'로 시작하면
# 엑셀이 수식(경우에 따라 DDE 외부 실행)으로 해석할 수 있다. 앞에 작은따옴표를 붙여
# 텍스트로 강제 — 값 자체는 안 바뀌고 엑셀에서 셀을 클릭하면 원문 그대로 보인다.
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _sanitize_cell(value):
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def generate_stats_excel() -> bytes:
    """선생님별 담당 원생 수 요약(표+막대그래프) + 전체 재적생 상세, 두 시트짜리 엑셀 파일을 바이트로 생성."""
    logger.info("[GENERATE_STATS_EXCEL] 시작")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "선생님별 원생 현황"
    ws1.append(["선생님", "파트", "담당 원생 수"])
    for row in db.get_teacher_student_counts():
        ws1.append([_sanitize_cell(row["display_name"]), _sanitize_cell(row["part"]), row["student_count"]])

    if ws1.max_row > 1:
        chart = BarChart()
        chart.title = "선생님별 담당 원생 수"
        chart.y_axis.title = "원생 수"
        data = Reference(ws1, min_col=3, min_row=1, max_row=ws1.max_row)
        cats = Reference(ws1, min_col=1, min_row=2, max_row=ws1.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws1.add_chart(chart, "E2")

    ws2 = wb.create_sheet("학생 상세")
    ws2.append(["이름", "파트", "나이", "MBTI", "가입 상태"])
    for row in db.get_all_active_students_for_export():
        ws2.append([
            _sanitize_cell(row["name"]), _sanitize_cell(row["instrument"]), row["age"], _sanitize_cell(row["mbti"] or "-"),
            "가입완료" if row["username"] else "미가입",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
